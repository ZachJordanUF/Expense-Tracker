from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
import calendar
import datetime

class Expense: # expense class 
    def __init__(self, name, amount, category):
        self.name = name
        self.amount = amount
        self.category = category

def main():
    print(f"🎯 Running Expense Tracker!")
    expense_file_path = "expenses.xlsx"  # Change the file extension to .xlsx
    budget = 2000

    # Get user input for expense.
    expense = get_user_expense() # Will return expense object from class

    # Write their expense to an Excel file.
    save_expense_to_excel(expense, expense_file_path)

    # Read file and summarize expenses.
    summarize_expenses(expense_file_path, budget)

def get_user_expense():
    print(f"🎯 Getting User Expense")
    expense_name = input("Enter expense name: ")
    expense_amount = float(input("Enter expense amount: "))
    expense_categories = [
        "🍔 Food",
        "🏠 Home",
        "💼 Work",
        "🎉 Fun",
        "✨ Misc",
    ]

    while True: # while user inputs name, category, and expense
        print("Select a category: ")
        for i, category_name in enumerate(expense_categories): # iterates through the index of expense_categories
            print(f"  {i + 1}. {category_name}")

        value_range = f"[1 - {len(expense_categories)}]"
        selected_index = int(input(f"Enter a category number {value_range}: ")) - 1

        if selected_index in range(len(expense_categories)):
            selected_category = expense_categories[selected_index] # category being selected by number
            new_expense = Expense( # creates object in class
                name=expense_name, category=selected_category, amount=expense_amount
            )
            return new_expense
        else:
            print("Invalid category. Please try again!")

def save_expense_to_excel(expense, expense_file_path):
    print(f"🎯 Saving User Expense: {expense.name}, {expense.amount}, {expense.category} to {expense_file_path}")
    wb = load_workbook(filename=expense_file_path) # We load an exel sheet where we are going to print out information  
    ws = wb.active # selects an active worksheet
    ws.append([expense.name, expense.amount, expense.category]) # adds a new row to the worksheet which is expense.name, expense.amount, expense.category
    wb.save(expense_file_path) # save the file 

def summarize_expenses(expense_file_path, budget):
    print(f"🎯 Summarizing User Expense")
    wb = load_workbook(filename=expense_file_path) # we load the workbook
    ws = wb.active # select the active sheet
    expenses = [] # create an empty list for expenses
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True): # iterates through the rows, starting from row 2
        expense_name, expense_amount, expense_category = row # populates these variables based on their corresponding cell in the row
        line_expense = Expense( 
            name=expense_name,
            amount=float(expense_amount),
            category=expense_category,
        )
        expenses.append(line_expense) # appends the object to the list

    amount_by_category = {} # creates a dictionary to store the amounts corresponding to the expense category 
    for expense in expenses:
        key = expense.category
        if key in amount_by_category: # if key exists add amount to existing category
            amount_by_category[key] += expense.amount
        else: # else create new amount for category
            amount_by_category[key] = expense.amount

    print("Expenses By Category 📈:")
    for key, amount in amount_by_category.items():
        print(f"  {key}: ${amount:.2f}")

    total_spent = sum([x.amount for x in expenses]) # sum up the total amounts of all expenses 
    print(f"💵 Total Spent: ${total_spent:.2f}")

    remaining_budget = budget - total_spent
    print(f"✅ Budget Remaining: ${remaining_budget:.2f}")

    now = datetime.datetime.now() # current date and time
    days_in_month = calendar.monthrange(now.year, now.month)[1]
    remaining_days = days_in_month - now.day

    daily_budget = remaining_budget / remaining_days # formula to calculate how much you can spend a month to stay within budget
    print(green(f"👉 Budget Per Day: ${daily_budget:.2f}"))

    # Create a bar chart (graph)
    chart = BarChart()
    chart.title = "Expenses by Category"
    data = Reference(ws, min_col=2, min_row=2, max_col=2, max_row=ws.max_row) # creates the data range for the bar chart. Specifies the column containing the expense amounts (column 2) and the rows where the data starts and ends (from row 2 to the last row with data)
    categories = Reference(ws, min_col=3, min_row=2, max_row=ws.max_row) # This creates a reference to the categories range for the bar chart. It specifies the column containing the expense categories (column 3) and the rows where the data starts and ends (from row 2 to the last row with data).
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, "E2")

    wb.save(expense_file_path)

def green(text):
    return f"\033[92m{text}\033[0m"

if __name__ == "__main__":
    main()
